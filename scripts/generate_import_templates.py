"""
Generate the three official Opriva Import Templates per OPRIVA_IMPORT_TEMPLATE_SPEC.md v2.0.

Outputs:
    public/templates/OPRIVA_IMPORT_TEMPLATE_MSP.xlsx
    public/templates/OPRIVA_IMPORT_TEMPLATE_INTERNAL_IT.xlsx
    public/templates/OPRIVA_IMPORT_TEMPLATE_CANONICAL.xlsx

Run:
    python3 scripts/generate_import_templates.py

No external network. Uses only the project-local Python stdlib + openpyxl.
All placeholder data is synthetic. No real customer data is embedded.
"""

from __future__ import annotations

import os
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

# -----------------------------------------------------------------------------
# Constants
# -----------------------------------------------------------------------------

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(REPO_ROOT, 'public', 'templates')
TEMPLATE_VERSION = '2.0'

# Opriva palette
NAVY = '0B1F3A'
TEAL = '0D9488'
WHITE = 'FFFFFF'
REQ_BG, REQ_TEXT = 'FEE2E2', 'B91C1C'
OPT_BG, OPT_TEXT = 'E2E8F0', '64748B'
SUG_BG, SUG_TEXT = 'FEF3C7', '92400E'
CALC_BG, CALC_TEXT = 'F1F5F9', '94A3B8'
ADV_BG, ADV_TEXT = 'EDE9FE', '7C3AED'
EXAMPLE_BG, EXAMPLE_TEXT = 'F8FAFC', '94A3B8'

# Sheet tab colors per category (spec §17.3)
TAB_INSTRUCTIONS = 'BFDBFE'
TAB_REFERENCE = '99F6E4'
TAB_RECORDS = 'BAE6FD'
TAB_COVERAGE = 'FEF3C7'
TAB_CONTRACTS = 'DDD6FE'
TAB_DOCUMENTS = 'E2E8F0'
TAB_TASKS = 'FBCFE8'
TAB_CUSTOM = 'FCE7F3'

# -----------------------------------------------------------------------------
# Controlled vocabulary (spec §16.4) — drives the hidden _Catalogs sheet
# -----------------------------------------------------------------------------

CATALOGS = {
    'RecordType': ['Client', 'Department'],
    'AlertPolicy': ['Workspace default', '90/60/30', '60/30/7', '30/7/1', 'Custom'],
    'DocumentType': ['Vendor Quote', 'Client Proposal', 'Purchase Order', 'Invoice',
                     'License Entitlement', 'Signed Contract', 'Warranty Document',
                     'Support Evidence', 'Compliance Evidence', 'Legal Document',
                     'Internal Memo', 'Other'],
    'ContractType': ['MSA', 'NDA', 'SLA', 'Service Contract', 'License Agreement',
                     'Hardware Contract', 'Other'],
    'CoverageKind': ['Warranty', 'Support', 'Maintenance'],
    'CoverageType': ['Manufacturer Warranty', 'Extended Warranty', 'Care Pack',
                     'SmartNet', 'Vendor Support', 'Managed Support',
                     'Subscription Support', 'Software Assurance', 'SLA Coverage',
                     'Maintenance Agreement', 'Other'],
    'SupportLevel': ['Bronze', 'Silver', 'Gold', 'Platinum', 'Standard',
                     'Premium', 'Mission Critical', 'Other'],
    'SuggestionBasis': ['file', 'inferred:purchase+term', 'inferred:license-term', 'manual'],
    'AssetType': ['Server', 'NAS / Storage', 'Firewall', 'Switch', 'Router',
                  'Laptop', 'Desktop', 'UPS', 'Other'],
    'EntitlementMetric': ['Devices', 'Users', 'Cores', 'VMs', 'Mailboxes',
                          'GB Storage', 'Sessions', 'Other'],
    'LicenseTerm': ['1 month', '1 year', '2 years', '3 years', '5 years',
                    'Perpetual', 'Custom'],
    'WarrantyTerm': ['1 year', '2 years', '3 years', '5 years', 'Custom'],
    'BillingCycle': ['Monthly', 'Annual', 'Multi-year', 'One-time', 'Other'],
    'Currency': ['USD', 'EUR', 'GBP', 'MXN', 'BRL', 'CRC', 'PAB', 'Other'],
    'Priority': ['Critical', 'High', 'Medium', 'Low'],
    'TaskStatus': ['Open', 'In Progress', 'Waiting', 'Done'],
    'TaskType': ['Quote Request', 'Document Request', 'Renewal Follow-up',
                 'Owner Assignment', 'Approval Request', 'Client Follow-up', 'Other'],
    'FieldType': ['Text', 'Number', 'Date', 'Dropdown', 'Currency', 'Checkbox',
                  'URL', 'Long Text'],
    'ApprovalStatus': ['Pending', 'Approved', 'Rejected', 'In Review'],
    'BusinessCriticality': ['Critical', 'High', 'Medium', 'Low'],
    'RelationshipType': ['Contains', 'Depends On', 'Replaces', 'Bundles With',
                         'Documents', 'Custom'],
    'Module': ['Licenses', 'Hardware', 'Coverage', 'Contracts', 'Documents',
               'Tasks', 'Clients / Departments', 'Renewal Packages'],
}

# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------

def fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type='solid')


def font_calibri(**kwargs) -> Font:
    kwargs.setdefault('name', 'Calibri')
    kwargs.setdefault('size', 11)
    return Font(**kwargs)


def add_catalogs_sheet(wb: Workbook) -> None:
    """Hidden _Catalogs sheet with one column per controlled vocabulary list."""
    ws = wb.create_sheet('_Catalogs')
    for col_idx, (name, values) in enumerate(CATALOGS.items(), start=1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.font = font_calibri(bold=True, color=NAVY)
        for row_idx, val in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col_idx, value=val)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(name) + 2)
    ws.sheet_state = 'hidden'


def catalog_range(name: str) -> str:
    """Return formula for a catalog dropdown referencing _Catalogs."""
    keys = list(CATALOGS.keys())
    col_idx = keys.index(name) + 1
    col_letter = get_column_letter(col_idx)
    n = len(CATALOGS[name])
    return f'=_Catalogs!${col_letter}$2:${col_letter}${n + 1}'


def style_header_row(ws, n_cols: int) -> None:
    f = fill(NAVY)
    ft = font_calibri(color=WHITE, bold=True, size=12)
    al = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = f
        cell.font = ft
        cell.alignment = al
    ws.row_dimensions[1].height = 28


def style_badge_row(ws, badges, n_cols: int) -> None:
    bg = {'REQ': REQ_BG, 'OPT': OPT_BG, 'SUG': SUG_BG, 'CALC': CALC_BG, 'ADV': ADV_BG}
    tx = {'REQ': REQ_TEXT, 'OPT': OPT_TEXT, 'SUG': SUG_TEXT, 'CALC': CALC_TEXT, 'ADV': ADV_TEXT}
    for c, b in enumerate(badges, start=1):
        cell = ws.cell(row=2, column=c, value=f'[{b}]')
        cell.fill = fill(bg.get(b, OPT_BG))
        cell.font = font_calibri(color=tx.get(b, OPT_TEXT), bold=True, italic=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18


def add_example_row(ws, example_values, n_cols: int) -> None:
    f = fill(EXAMPLE_BG)
    ft = font_calibri(color=EXAMPLE_TEXT, italic=True)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=3, column=c)
        v = example_values[c - 1] if c <= len(example_values) else ''
        if c == 1:
            v = f'DELETE BEFORE IMPORT — {v}' if v else 'DELETE BEFORE IMPORT'
        cell.value = v
        cell.fill = f
        cell.font = ft
        cell.alignment = Alignment(horizontal='left', vertical='center')


def setup_sheet(
    wb: Workbook,
    sheet_name: str,
    columns: list,
    badges: list,
    example: list,
    tab_color: str = None,
    freeze_col_a: bool = False,
    comments: dict = None,
    dropdowns: dict = None,
    calc_cols: list = None,
):
    """
    Build a data sheet with header + badge + example + freeze + filter + dropdowns.

    columns: list of column header names
    badges: same length list of REQ/OPT/SUG/CALC/ADV markers
    example: same length list of example values
    comments: dict {col_idx_1based: comment_text}
    dropdowns: dict {col_idx_1based: catalog_name}
    calc_cols: list of 1-based col indices to style as Calculated
    """
    ws = wb.create_sheet(sheet_name)
    n_cols = len(columns)
    assert len(badges) == n_cols, f'{sheet_name}: badges len mismatch'

    for c, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        if comments and c in comments:
            cell.comment = Comment(comments[c], 'Opriva Template')

    style_header_row(ws, n_cols)
    style_badge_row(ws, badges, n_cols)
    add_example_row(ws, example, n_cols)

    for c in range(1, n_cols + 1):
        col_letter = get_column_letter(c)
        h = len(columns[c - 1])
        ws.column_dimensions[col_letter].width = max(18, min(38, h + 4))

    if freeze_col_a:
        ws.freeze_panes = 'B4'
    else:
        ws.freeze_panes = 'A4'

    last = get_column_letter(n_cols)
    ws.auto_filter.ref = f'A1:{last}1'

    if tab_color:
        ws.sheet_properties.tabColor = tab_color

    if dropdowns:
        for col_idx, cat_name in dropdowns.items():
            col_letter = get_column_letter(col_idx)
            formula = catalog_range(cat_name)
            dv = DataValidation(type='list', formula1=formula, allow_blank=True)
            dv.errorTitle = 'Invalid value'
            dv.error = f'Value must be one of the allowed {cat_name} options.'
            dv.promptTitle = cat_name
            dv.prompt = f'Pick a {cat_name} value from the dropdown.'
            dv.add(f'{col_letter}4:{col_letter}998')
            ws.add_data_validation(dv)

    if calc_cols:
        f = fill(CALC_BG)
        ft = font_calibri(color=CALC_TEXT, italic=True)
        for col_idx in calc_cols:
            col_letter = get_column_letter(col_idx)
            for row in range(4, 999):
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = f
                cell.font = ft

    return ws


def build_instructions_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    mode_label: str,
    fill_order: list,
    template_marker: str,
    tab_color: str,
):
    """Insert Instructions sheet at index 0 with version marker and all guidance."""
    ws = wb.create_sheet(sheet_name, 0)

    ws['A1'] = template_marker
    ws['A1'].font = font_calibri(bold=True, size=10, color='64748B')

    ws['A3'] = title
    ws['A3'].font = font_calibri(bold=True, size=18, color=NAVY)

    ws['A4'] = f'Workspace mode: {mode_label}'
    ws['A4'].font = font_calibri(bold=True, size=12, color=TEAL)

    ws['A5'] = (f'Template version: {TEMPLATE_VERSION}    |    '
                f'See OPRIVA_IMPORT_TEMPLATE_SPEC.md v{TEMPLATE_VERSION} in the Opriva repository.')
    ws['A5'].font = font_calibri(size=10, color='64748B')

    row = 7
    ws.cell(row=row, column=1, value='Fill order').font = font_calibri(bold=True, size=14, color=NAVY)
    row += 1
    for i, step in enumerate(fill_order, start=1):
        ws.cell(row=row, column=1, value=f'{i}. {step}').font = font_calibri()
        row += 1
    row += 1

    ws.cell(row=row, column=1, value='Column badges').font = font_calibri(bold=True, size=14, color=NAVY)
    row += 1
    legend = [
        ('[REQ]', 'Required — row cannot be imported without this value', REQ_BG, REQ_TEXT),
        ('[OPT]', 'Optional — can be completed later in the record drawer', OPT_BG, OPT_TEXT),
        ('[SUG]', 'Suggested — Opriva infers; user approves in the import preview', SUG_BG, SUG_TEXT),
        ('[CALC]', 'Calculated — derived by Opriva; do not fill', CALC_BG, CALC_TEXT),
        ('[ADV]', 'Advanced — Canonical template only, expert column', ADV_BG, ADV_TEXT),
    ]
    for code, desc, bg, tx in legend:
        c1 = ws.cell(row=row, column=1, value=code)
        c1.fill = fill(bg)
        c1.font = font_calibri(color=tx, bold=True, italic=True, size=10)
        c1.alignment = Alignment(horizontal='center', vertical='center')
        c2 = ws.cell(row=row, column=2, value=desc)
        c2.font = font_calibri()
        row += 1
    row += 1

    ws.cell(row=row, column=1, value='Date format').font = font_calibri(bold=True, size=14, color=NAVY)
    row += 1
    ws.cell(row=row, column=1, value='All dates must be in YYYY-MM-DD (ISO 8601). Examples: 2026-05-31, 2027-01-15.').font = font_calibri()
    row += 1
    ws.cell(row=row, column=1, value='Avoid 5/31/26, 31-05-2026, or May 31 2026.').font = font_calibri(italic=True, color='64748B')
    row += 2

    ws.cell(row=row, column=1, value='Cross-sheet references').font = font_calibri(bold=True, size=14, color=NAVY)
    row += 1
    refs = [
        'Client / Department Name — used by every sheet that references a client or department',
        'Package Reference — links Licenses, Hardware, Coverage, Contracts, Documents to a parent Renewal Package',
        'Covered Record Reference — links Coverage rows to a License or Hardware record',
        'Linked Record Reference — links Documents and Tasks to a specific record in any module',
    ]
    for r in refs:
        ws.cell(row=row, column=1, value=f'• {r}').font = font_calibri()
        row += 1
    row += 1

    ws.cell(row=row, column=1, value='Calculated fields').font = font_calibri(bold=True, size=14, color=NAVY)
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=('Do not fill: System Status, Days to Expiration, Margin, Risk, '
               'Renewal Stage, Validity Status, Missing Evidence, Alert Status. '
               'Opriva derives these from other values.'),
    ).font = font_calibri(italic=True, color=CALC_TEXT)
    row += 2

    ws.cell(row=row, column=1, value='Coverage (Warranty / Support / Maintenance)').font = font_calibri(bold=True, size=14, color=NAVY)
    row += 1
    cov = [
        'Each warranty, support and maintenance coverage is its own row on the Coverage sheet.',
        'Link a coverage to its parent License or Hardware via Covered Record Reference.',
        'Coverage Kind discriminates: Warranty / Support / Maintenance.',
        'Coverage Type catalog includes Manufacturer Warranty, Extended Warranty, Care Pack, SmartNet, Vendor Support, Managed Support, Subscription Support, Software Assurance, SLA Coverage, Maintenance Agreement, Other.',
        'Opriva can infer coverage from Purchase Date + Warranty Term (Hardware) or License Start + Expiration + Support keyword (License). Inferred coverages appear as Suggested in the import preview and require user approval before creation.',
        'Maintenance coverage is never inferred — provide explicitly.',
        'File-provided coverage values always take precedence over Opriva inferences.',
    ]
    for n in cov:
        ws.cell(row=row, column=1, value=f'• {n}').font = font_calibri()
        row += 1
    row += 1

    ws.cell(row=row, column=1, value='Documents — metadata only').font = font_calibri(bold=True, size=14, color=NAVY)
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=('The Documents sheet imports document metadata only. Actual files must be uploaded '
               'separately via the Opriva Documents module or the Documents tab of the linked '
               'record drawer after import.'),
    ).font = font_calibri()
    row += 2

    ws.cell(row=row, column=1, value='Security and sensitivity').font = font_calibri(bold=True, size=14, color='B91C1C')
    row += 1
    sec = [
        'Contact Name and Contact Email contain personal data (PII). Avoid uploading real contact data in sandbox mode.',
        'Vendor Cost, Sale Price, Annual Value and Annual Cost are commercial-sensitive. Backend will enforce permission boundaries.',
        'Serial Numbers are asset-sensitive; ensure uniqueness.',
        'File Name column is informational. Actual document files must be uploaded separately via the Documents module.',
        'Do not store passwords, API keys, license activation keys, internal IP addresses, tax IDs / SSN, bank info or health data in any sheet — including Custom Fields.',
    ]
    for n in sec:
        ws.cell(row=row, column=1, value=f'! {n}').font = font_calibri(color='7F1D1D')
        row += 1
    row += 1

    ws.cell(row=row, column=1, value='When to use Path A instead (AI-assisted guided mapping)').font = font_calibri(bold=True, size=14, color=NAVY)
    row += 1
    path_a = [
        'If you have an existing vendor export (Microsoft CSP, Veeam, Fortinet, VMware, Cisco SmartNet, Dell warranty extract, HPE Care Pack, QNAP, distributor monthly report), use Data Import → Upload file (Path A AI-assisted mapping) instead of copying values into this template.',
        'See IMPORT_MAPPING_*.md docs in the Opriva repository for vendor-specific mapping references.',
    ]
    for n in path_a:
        ws.cell(row=row, column=1, value=n).font = font_calibri()
        row += 1

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 110
    ws.sheet_properties.tabColor = tab_color
    return ws


# -----------------------------------------------------------------------------
# Column definitions per sheet, per template
# -----------------------------------------------------------------------------

def add_coverage_sheet(wb: Workbook, value_label: str):
    """Coverage sheet (shared shape, only value column label differs per workspace)."""
    cols = [
        'Coverage Reference', 'Coverage Kind', 'Covered Record Reference', 'Coverage Type',
        'Provider', 'Support Level', 'Start Date', 'End Date', 'Support Reference',
        'Alert Policy', 'Owner', value_label, 'Currency', 'Suggestion Basis', 'Notes',
    ]
    badges = ['REQ', 'REQ', 'REQ', 'REQ',
              'OPT', 'OPT', 'OPT', 'REQ', 'OPT',
              'OPT', 'OPT', 'OPT', 'OPT', 'OPT', 'OPT']
    example = [
        'COV-001', 'Warranty', 'HW-001', 'Manufacturer Warranty',
        'Dell Direct', 'Standard', '2025-08-15', '2028-08-15', 'PO-7842',
        'Workspace default', 'It Owner A', '0', 'USD', 'file', 'sample row',
    ]
    dropdowns = {2: 'CoverageKind', 4: 'CoverageType', 6: 'SupportLevel',
                 10: 'AlertPolicy', 13: 'Currency', 14: 'SuggestionBasis'}
    comments = {
        2: 'Warranty / Support / Maintenance — discriminator for the related coverage record.',
        3: 'Must match an existing License Reference or Hardware Reference from those sheets.',
        4: 'Specific coverage type. See spec §9.3 for catalog.',
        7: 'Coverage start date (YYYY-MM-DD). May be inferred from parent record.',
        8: 'Coverage end date (YYYY-MM-DD). Required — drives status and alerts.',
        14: 'Origin of this coverage row: file (default), inferred:purchase+term, inferred:license-term, or manual.',
    }
    setup_sheet(wb, 'Coverage', cols, badges, example,
                tab_color=TAB_COVERAGE, freeze_col_a=True,
                comments=comments, dropdowns=dropdowns)


def add_documents_sheet(wb: Workbook):
    cols = [
        'Document Reference', 'Package Reference', 'Linked Record Reference',
        'Document Name', 'Document Type', 'File Name', 'Uploaded By', 'Notes',
    ]
    badges = ['REQ', 'OPT', 'OPT', 'REQ', 'REQ', 'OPT', 'OPT', 'OPT']
    example = [
        'DOC-001', 'PKG-001', 'LIC-001',
        'License entitlement certificate', 'License Entitlement',
        'license-entitlement-2026.pdf', 'Owner A', 'sample row',
    ]
    dropdowns = {5: 'DocumentType'}
    comments = {
        5: 'Document Type controlled catalog. See spec §11.2.',
        6: 'INFORMATIONAL ONLY. Actual files must be uploaded separately via the Documents module — this sheet does not upload files.',
    }
    setup_sheet(wb, 'Documents', cols, badges, example,
                tab_color=TAB_DOCUMENTS, comments=comments, dropdowns=dropdowns)


def build_msp_template():
    wb = Workbook()
    wb.remove(wb.active)

    build_instructions_sheet(
        wb,
        sheet_name='Instructions_MSP',
        title='Opriva Import Template — MSP / Integrator',
        mode_label='MSP / Integrator',
        fill_order=[
            'Fill the Clients sheet first — all other sheets reference these names.',
            'Fill Renewal Packages if you group licenses, hardware and coverage by deal.',
            'Fill Licenses for software subscriptions and licenses.',
            'Fill Hardware for physical IT assets.',
            'Fill Coverage for warranty, support, maintenance, SLA, Software Assurance, Subscription Support, Care Pack or SmartNet coverage linked to License or Hardware records.',
            'Fill Contracts for commercial agreements (MSA, NDA, SLA, Service Contract, License Agreement) — not support coverage.',
            'Fill Documents for document metadata (files uploaded separately via the Documents module).',
            'Delete the example row (row 3) in every sheet you fill before uploading.',
            'Upload via Data Import → Upload Opriva Template.',
        ],
        template_marker='OPRIVA_TEMPLATE_V2.0_MSP',
        tab_color=TAB_INSTRUCTIONS,
    )

    # Clients sheet
    setup_sheet(
        wb, 'Clients',
        columns=['Client Reference', 'Client Name', 'Country', 'Account Owner',
                 'Contact Name', 'Contact Email', 'Notes'],
        badges=['REQ', 'REQ', 'OPT', 'OPT', 'OPT', 'OPT', 'OPT'],
        example=['CL-001', 'Acme Holdings', 'PA', 'Account Owner A',
                 'Contact A', 'contact@example.invalid', 'sample row'],
        tab_color=TAB_REFERENCE,
        comments={5: 'PII — see Instructions sheet security section.',
                  6: 'PII — see Instructions sheet security section.'},
    )

    # Renewal Packages — MSP
    setup_sheet(
        wb, 'Renewal_Packages',
        columns=['Package Reference', 'Package Name', 'Client', 'Brand / Vendor',
                 'Distributor', 'Reseller / Partner', 'PO / Order Reference',
                 'Quote Reference', 'Invoice Reference', 'Start Date',
                 'Expiration / Renewal Date', 'Sale Price / Annual Value', 'Vendor Cost',
                 'Currency', 'Account Owner', 'Alert Policy', 'Notes'],
        badges=['REQ', 'REQ', 'REQ', 'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT'],
        example=['PKG-001', 'Acme — 2026 Renewal Bundle', 'Acme Holdings', 'Vendor Brand A',
                 'Distributor A', 'Reseller B', 'PO-7842',
                 'Q-5512', 'INV-1102', '2026-01-01',
                 '2027-01-01', '50000', '38000',
                 'USD', 'Account Owner A', 'Workspace default', 'sample row'],
        tab_color=TAB_REFERENCE,
        dropdowns={14: 'Currency', 16: 'AlertPolicy'},
    )

    # Licenses — MSP
    setup_sheet(
        wb, 'Licenses',
        columns=['License Reference', 'Package Reference', 'Client', 'License / Product',
                 'Brand', 'Distributor', 'Reseller / Partner', 'Quantity', 'Entitlement Metric',
                 'Start Date', 'Expiration / Renewal Date', 'License Term', 'Billing Cycle',
                 'Alert Policy', 'Renewal Owner', 'Sale Price / Annual Value', 'Vendor Cost',
                 'Currency', 'PO / Order Reference', 'Invoice Reference', 'Invoice Date',
                 'Source Reference', 'Source Status', 'Notes'],
        badges=['REQ', 'OPT', 'REQ', 'REQ',
                'OPT', 'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'REQ', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT'],
        example=['LIC-001', 'PKG-001', 'Acme Holdings', 'Sample Subscription Plan',
                 'Vendor Brand A', 'Distributor A', 'Reseller B', '50', 'Users',
                 '2026-01-17', '2027-01-17', '1 year', 'Annual',
                 'Workspace default', 'Renewal Owner A', '12500', '8900',
                 'USD', 'PO-7842', 'INV-1102', '2026-01-15',
                 'VND-REF-001', 'Active', 'sample row'],
        tab_color=TAB_RECORDS,
        freeze_col_a=True,
        dropdowns={9: 'EntitlementMetric', 12: 'LicenseTerm', 13: 'BillingCycle',
                   14: 'AlertPolicy', 18: 'Currency'},
    )

    # Hardware — MSP
    setup_sheet(
        wb, 'Hardware',
        columns=['Hardware Reference', 'Package Reference', 'Client', 'Asset Name',
                 'Asset Type', 'Brand', 'Model', 'Serial Number', 'Distributor',
                 'Reseller / Partner', 'Purchase Date', 'Warranty Term', 'Asset Value',
                 'Currency', 'Location', 'Owner', 'Notes'],
        badges=['REQ', 'OPT', 'REQ', 'REQ',
                'OPT', 'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT'],
        example=['HW-001', 'PKG-001', 'Acme Holdings', 'Sample Server R750',
                 'Server', 'Hardware Brand B', 'R750', 'SN-XXXX-0001', 'Distributor A',
                 'Reseller B', '2025-08-15', '3 years', '12500',
                 'USD', 'Site HQ', 'Owner A', 'sample row'],
        tab_color=TAB_RECORDS,
        freeze_col_a=True,
        dropdowns={5: 'AssetType', 12: 'WarrantyTerm', 14: 'Currency'},
        comments={11: 'Purchase Date + Warranty Term lets Opriva suggest a Warranty Coverage row at import. Review and approve in the import preview.',
                  12: 'Warranty Term used to infer Warranty End Date when Coverage sheet is empty.'},
    )

    add_coverage_sheet(wb, value_label='Annual Value')

    # Contracts — MSP
    setup_sheet(
        wb, 'Contracts',
        columns=['Contract Reference', 'Package Reference', 'Contract Name', 'Contract Type',
                 'Client', 'Counterparty', 'Start Date', 'End Date', 'Owner',
                 'Alert Policy', 'Annual Value', 'Currency', 'Notice Period', 'Notes'],
        badges=['REQ', 'OPT', 'REQ', 'REQ',
                'REQ', 'OPT', 'OPT', 'REQ', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT', 'OPT'],
        example=['CTR-001', 'PKG-001', 'Master Services Agreement — Acme', 'MSA',
                 'Acme Holdings', 'Counterparty Inc.', '2025-01-01', '2028-01-01', 'Owner A',
                 'Workspace default', '0', 'USD', '60 days', 'sample row'],
        tab_color=TAB_CONTRACTS,
        dropdowns={4: 'ContractType', 10: 'AlertPolicy', 12: 'Currency'},
    )

    add_documents_sheet(wb)

    add_catalogs_sheet(wb)

    out_path = os.path.join(OUT_DIR, 'OPRIVA_IMPORT_TEMPLATE_MSP.xlsx')
    wb.save(out_path)
    return out_path


def build_internal_it_template():
    wb = Workbook()
    wb.remove(wb.active)

    build_instructions_sheet(
        wb,
        sheet_name='Instructions_Internal_IT',
        title='Opriva Import Template — Internal IT',
        mode_label='Internal IT',
        fill_order=[
            'Fill the Departments sheet first — all other sheets reference these names.',
            'Fill Renewal Packages if your IT organization groups by package (optional).',
            'Fill Licenses for software subscriptions and licenses with annual cost and approval status.',
            'Fill Hardware for physical IT assets with custodian, location and cost center.',
            'Fill Coverage for warranty, support, maintenance, SLA, Software Assurance, Subscription Support, Care Pack or SmartNet coverage linked to License or Hardware records.',
            'Fill Contracts for commercial agreements (MSA, NDA, SLA, Service Contract, License Agreement) — not support coverage.',
            'Fill Documents for document metadata (files uploaded separately via the Documents module).',
            'Delete the example row (row 3) in every sheet you fill before uploading.',
            'Upload via Data Import → Upload Opriva Template.',
        ],
        template_marker='OPRIVA_TEMPLATE_V2.0_INTERNAL_IT',
        tab_color=TAB_INSTRUCTIONS,
    )

    # Departments
    setup_sheet(
        wb, 'Departments',
        columns=['Department Reference', 'Department Name', 'Business Unit', 'Cost Center',
                 'Location', 'IT Owner', 'Budget Owner', 'Notes'],
        badges=['REQ', 'REQ', 'OPT', 'OPT', 'OPT', 'OPT', 'OPT', 'OPT'],
        example=['DPT-001', 'Finance', 'Corporate Services', 'CC-FIN-01',
                 'HQ Floor 3', 'IT Owner A', 'Budget Owner B', 'sample row'],
        tab_color=TAB_REFERENCE,
    )

    # Renewal Packages — IT (optional)
    setup_sheet(
        wb, 'Renewal_Packages',
        columns=['Package Reference', 'Package Name', 'Department', 'Brand / Vendor',
                 'Provider / Vendor', 'PO / Order Reference', 'Invoice Reference',
                 'Start Date', 'Expiration / Renewal Date', 'Annual Cost', 'Currency',
                 'Budget Owner', 'IT Owner', 'Cost Center', 'Approval Status',
                 'Alert Policy', 'Notes'],
        badges=['REQ', 'REQ', 'REQ', 'OPT',
                'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT'],
        example=['PKG-001', 'Finance — Microsoft 365 2026', 'Finance', 'Vendor Brand A',
                 'Provider A', 'PO-9921', 'INV-2202',
                 '2026-01-01', '2027-01-01', '16200', 'USD',
                 'Budget Owner B', 'IT Owner A', 'CC-FIN-01', 'Approved',
                 'Workspace default', 'sample row'],
        tab_color=TAB_REFERENCE,
        dropdowns={11: 'Currency', 15: 'ApprovalStatus', 16: 'AlertPolicy'},
    )

    # Licenses — IT
    setup_sheet(
        wb, 'Licenses',
        columns=['License Reference', 'Package Reference', 'Department', 'License / Product',
                 'Brand', 'Provider / Vendor', 'Quantity', 'Entitlement Metric',
                 'Start Date', 'Expiration / Renewal Date', 'License Term', 'Billing Cycle',
                 'Alert Policy', 'IT Owner', 'Budget Owner', 'Annual Cost', 'Currency',
                 'PO / Order Reference', 'Invoice Reference', 'Approval Status',
                 'Business Criticality', 'Cost Center', 'Source Reference', 'Source Status',
                 'Notes'],
        badges=['REQ', 'OPT', 'REQ', 'REQ',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'REQ', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT'],
        example=['LIC-001', 'PKG-001', 'Finance', 'Sample Software Plan',
                 'Vendor Brand A', 'Provider A', '45', 'Users',
                 '2026-01-01', '2027-01-01', '1 year', 'Annual',
                 'Workspace default', 'IT Owner A', 'Budget Owner B', '16200', 'USD',
                 'PO-9921', 'INV-2202', 'Approved',
                 'High', 'CC-FIN-01', 'VND-REF-001', 'Active',
                 'sample row'],
        tab_color=TAB_RECORDS,
        freeze_col_a=True,
        dropdowns={8: 'EntitlementMetric', 11: 'LicenseTerm', 12: 'BillingCycle',
                   13: 'AlertPolicy', 17: 'Currency', 20: 'ApprovalStatus',
                   21: 'BusinessCriticality'},
    )

    # Hardware — IT
    setup_sheet(
        wb, 'Hardware',
        columns=['Hardware Reference', 'Package Reference', 'Department', 'Asset Name',
                 'Asset Type', 'Brand', 'Model', 'Serial Number', 'Provider / Vendor',
                 'Purchase Date', 'Warranty Term', 'Asset Value', 'Currency',
                 'Location', 'Asset Custodian', 'Budget Owner', 'Cost Center',
                 'Approval Status', 'Business Criticality', 'Notes'],
        badges=['REQ', 'OPT', 'REQ', 'REQ',
                'OPT', 'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT'],
        example=['HW-001', 'PKG-001', 'Engineering', 'Sample Server R750',
                 'Server', 'Hardware Brand B', 'R750', 'SN-XXXX-0001', 'Provider A',
                 '2025-08-15', '3 years', '12500', 'USD',
                 'Site HQ Rack 3', 'Custodian Eng Ops', 'Budget Owner B', 'CC-ENG-02',
                 'Approved', 'High', 'sample row'],
        tab_color=TAB_RECORDS,
        freeze_col_a=True,
        dropdowns={5: 'AssetType', 11: 'WarrantyTerm', 13: 'Currency',
                   18: 'ApprovalStatus', 19: 'BusinessCriticality'},
        comments={10: 'Purchase Date + Warranty Term lets Opriva suggest a Warranty Coverage row at import. Review and approve in the import preview.',
                  11: 'Warranty Term used to infer Warranty End Date when Coverage sheet is empty.'},
    )

    add_coverage_sheet(wb, value_label='Annual Cost')

    # Contracts — IT
    setup_sheet(
        wb, 'Contracts',
        columns=['Contract Reference', 'Package Reference', 'Contract Name', 'Contract Type',
                 'Department', 'Counterparty', 'Start Date', 'End Date', 'IT Owner',
                 'Budget Owner', 'Alert Policy', 'Annual Cost', 'Currency',
                 'Approval Status', 'Notice Period', 'Notes'],
        badges=['REQ', 'OPT', 'REQ', 'REQ',
                'REQ', 'OPT', 'OPT', 'REQ', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT'],
        example=['CTR-001', 'PKG-001', 'Master Services Agreement — Finance', 'MSA',
                 'Finance', 'Counterparty Inc.', '2025-01-01', '2028-01-01', 'IT Owner A',
                 'Budget Owner B', 'Workspace default', '0', 'USD',
                 'Approved', '60 days', 'sample row'],
        tab_color=TAB_CONTRACTS,
        dropdowns={4: 'ContractType', 11: 'AlertPolicy', 13: 'Currency', 14: 'ApprovalStatus'},
    )

    add_documents_sheet(wb)

    add_catalogs_sheet(wb)

    out_path = os.path.join(OUT_DIR, 'OPRIVA_IMPORT_TEMPLATE_INTERNAL_IT.xlsx')
    wb.save(out_path)
    return out_path


def build_canonical_template():
    wb = Workbook()
    wb.remove(wb.active)

    build_instructions_sheet(
        wb,
        sheet_name='Instructions_Canonical',
        title='Opriva Import Template — Canonical / Advanced',
        mode_label='Canonical / Advanced (covers MSP, Internal IT, Hybrid)',
        fill_order=[
            'Fill the Clients_Departments sheet first. Use Record Type = Client (MSP) or Department (Internal IT).',
            'Fill Renewal Packages if you group records by deal or renewal cycle.',
            'Fill Licenses, Hardware, Coverage, Contracts, Documents as applicable.',
            'Fill Tasks for pre-loaded follow-up actions linked to records.',
            'Fill Custom Fields for workspace-specific extra columns that do not fit any standard field.',
            'Fill Relationships for advanced cross-record links beyond Package and Covered Record references.',
            'Delete the example row (row 3) in every sheet you fill before uploading.',
            'Upload via Data Import → Upload Opriva Template.',
        ],
        template_marker='OPRIVA_TEMPLATE_V2.0_CANONICAL',
        tab_color=TAB_INSTRUCTIONS,
    )

    # Clients_Departments
    setup_sheet(
        wb, 'Clients_Departments',
        columns=['Record Type', 'Reference', 'Name', 'Business Unit', 'Cost Center',
                 'Country', 'Location', 'Owner', 'Budget Owner',
                 'Contact Name', 'Contact Email', 'Notes'],
        badges=['REQ', 'REQ', 'REQ', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT'],
        example=['Client', 'CL-001', 'Acme Holdings', '', '',
                 'PA', '', 'Account Owner A', '',
                 'Contact A', 'contact@example.invalid', 'sample row'],
        tab_color=TAB_REFERENCE,
        dropdowns={1: 'RecordType'},
        comments={10: 'PII — see Instructions sheet security section.',
                  11: 'PII — see Instructions sheet security section.'},
    )

    # Renewal Packages — Canonical
    setup_sheet(
        wb, 'Renewal_Packages',
        columns=['Package Reference', 'Package Name', 'Client / Department', 'Brand / Vendor',
                 'Provider / Distributor', 'Reseller / Partner', 'PO / Order Reference',
                 'Quote Reference', 'Invoice Reference', 'Start Date',
                 'Expiration / Renewal Date', 'Sale Price / Annual Value', 'Vendor Cost',
                 'Annual Cost', 'Currency', 'Owner', 'Budget Owner', 'Cost Center',
                 'Approval Status', 'Alert Policy', 'Notes'],
        badges=['REQ', 'REQ', 'REQ', 'OPT',
                'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT'],
        example=['PKG-001', 'Acme — 2026 Renewal Bundle', 'Acme Holdings', 'Vendor Brand A',
                 'Distributor A', 'Reseller B', 'PO-7842',
                 'Q-5512', 'INV-1102', '2026-01-01',
                 '2027-01-01', '50000', '38000',
                 '', 'USD', 'Account Owner A', '', '',
                 '', 'Workspace default', 'sample row'],
        tab_color=TAB_REFERENCE,
        dropdowns={15: 'Currency', 19: 'ApprovalStatus', 20: 'AlertPolicy'},
    )

    # Licenses — Canonical (MSP+IT merged + Calculated)
    setup_sheet(
        wb, 'Licenses',
        columns=['License Reference', 'Package Reference', 'Client / Department', 'License / Product',
                 'Brand', 'Distributor', 'Reseller / Partner', 'Provider / Vendor',
                 'Quantity', 'Entitlement Metric', 'Start Date', 'Expiration / Renewal Date',
                 'License Term', 'Billing Cycle', 'Alert Policy', 'Owner',
                 'Budget Owner', 'Sale Price / Annual Value', 'Vendor Cost', 'Annual Cost',
                 'Currency', 'PO / Order Reference', 'Invoice Reference', 'Invoice Date',
                 'Approval Status', 'Business Criticality', 'Cost Center',
                 'Source Reference', 'Source Status', 'Notes',
                 'System Status', 'Days to Expiration', 'Margin $', 'Margin %'],
        badges=['REQ', 'OPT', 'REQ', 'REQ',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'REQ',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT',
                'CALC', 'CALC', 'CALC', 'CALC'],
        example=['LIC-001', 'PKG-001', 'Acme Holdings', 'Sample Subscription Plan',
                 'Vendor Brand A', 'Distributor A', 'Reseller B', '',
                 '50', 'Users', '2026-01-17', '2027-01-17',
                 '1 year', 'Annual', 'Workspace default', 'Renewal Owner A',
                 '', '12500', '8900', '',
                 'USD', 'PO-7842', 'INV-1102', '2026-01-15',
                 '', '', '',
                 'VND-REF-001', 'Active', 'sample row',
                 '', '', '', ''],
        tab_color=TAB_RECORDS,
        freeze_col_a=True,
        dropdowns={10: 'EntitlementMetric', 13: 'LicenseTerm', 14: 'BillingCycle',
                   15: 'AlertPolicy', 21: 'Currency', 25: 'ApprovalStatus',
                   26: 'BusinessCriticality'},
        calc_cols=[31, 32, 33, 34],
        comments={31: 'CALCULATED — do not fill. Opriva derives from Expiration Date + Alert Policy.',
                  32: 'CALCULATED — do not fill. Days from today to Expiration / Renewal Date.',
                  33: 'CALCULATED — do not fill. Sale Price minus Vendor Cost. MSP only.',
                  34: 'CALCULATED — do not fill. Margin dollars divided by Sale Price. MSP only.'},
    )

    # Hardware — Canonical
    setup_sheet(
        wb, 'Hardware',
        columns=['Hardware Reference', 'Package Reference', 'Client / Department', 'Asset Name',
                 'Asset Type', 'Brand', 'Model', 'Serial Number',
                 'Distributor', 'Reseller / Partner', 'Provider / Vendor',
                 'Purchase Date', 'Warranty Term', 'Asset Value', 'Currency',
                 'Location', 'Owner', 'Asset Custodian', 'Budget Owner', 'Cost Center',
                 'Approval Status', 'Business Criticality', 'Notes',
                 'Days to Warranty End', 'System Status'],
        badges=['REQ', 'OPT', 'REQ', 'REQ',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT',
                'CALC', 'CALC'],
        example=['HW-001', 'PKG-001', 'Engineering', 'Sample Server R750',
                 'Server', 'Hardware Brand B', 'R750', 'SN-XXXX-0001',
                 'Distributor A', 'Reseller B', '',
                 '2025-08-15', '3 years', '12500', 'USD',
                 'Site HQ Rack 3', 'Owner A', '', '', '',
                 '', '', 'sample row',
                 '', ''],
        tab_color=TAB_RECORDS,
        freeze_col_a=True,
        dropdowns={5: 'AssetType', 13: 'WarrantyTerm', 15: 'Currency',
                   21: 'ApprovalStatus', 22: 'BusinessCriticality'},
        calc_cols=[24, 25],
        comments={12: 'Purchase Date + Warranty Term lets Opriva suggest a Warranty Coverage row at import.',
                  24: 'CALCULATED — do not fill.',
                  25: 'CALCULATED — do not fill.'},
    )

    add_coverage_sheet(wb, value_label='Annual Value / Cost')

    # Contracts — Canonical
    setup_sheet(
        wb, 'Contracts',
        columns=['Contract Reference', 'Package Reference', 'Contract Name', 'Contract Type',
                 'Client / Department', 'Counterparty', 'Start Date', 'End Date',
                 'Owner', 'Budget Owner', 'Alert Policy', 'Annual Value', 'Annual Cost',
                 'Currency', 'Approval Status', 'Notice Period', 'Notes'],
        badges=['REQ', 'OPT', 'REQ', 'REQ',
                'REQ', 'OPT', 'OPT', 'REQ',
                'OPT', 'OPT', 'OPT', 'OPT', 'OPT',
                'OPT', 'OPT', 'OPT', 'OPT'],
        example=['CTR-001', 'PKG-001', 'Master Services Agreement', 'MSA',
                 'Acme Holdings', 'Counterparty Inc.', '2025-01-01', '2028-01-01',
                 'Owner A', '', 'Workspace default', '0', '',
                 'USD', '', '60 days', 'sample row'],
        tab_color=TAB_CONTRACTS,
        dropdowns={4: 'ContractType', 11: 'AlertPolicy', 14: 'Currency', 15: 'ApprovalStatus'},
    )

    add_documents_sheet(wb)

    # Tasks
    setup_sheet(
        wb, 'Tasks',
        columns=['Task Reference', 'Linked Record Reference', 'Task Title', 'Task Type',
                 'Owner', 'Due Date', 'Priority', 'Status', 'Notes'],
        badges=['REQ', 'OPT', 'REQ', 'OPT', 'OPT', 'OPT', 'OPT', 'OPT', 'OPT'],
        example=['TSK-001', 'LIC-001', 'Request renewal quote', 'Quote Request',
                 'Owner A', '2026-11-01', 'High', 'Open', 'sample row'],
        tab_color=TAB_TASKS,
        dropdowns={4: 'TaskType', 7: 'Priority', 8: 'TaskStatus'},
    )

    # Custom Fields
    setup_sheet(
        wb, 'Custom_Fields',
        columns=['Module', 'Record Reference', 'Field Name', 'Field Type', 'Value', 'Notes'],
        badges=['REQ', 'REQ', 'REQ', 'REQ', 'REQ', 'OPT'],
        example=['Licenses', 'LIC-001', 'TM Program Number', 'Text', 'TMP-12345', 'sample row'],
        tab_color=TAB_CUSTOM,
        dropdowns={1: 'Module', 4: 'FieldType'},
        comments={3: 'Do not duplicate any standard column name. See Instructions sheet for safety rules.',
                  5: 'Do not store passwords, API keys, IP addresses, tax IDs, bank info, or other sensitive credentials.'},
    )

    # Relationships
    setup_sheet(
        wb, 'Relationships',
        columns=['Source Module', 'Source Reference', 'Relationship Type', 'Target Module',
                 'Target Reference', 'Notes'],
        badges=['REQ', 'REQ', 'REQ', 'REQ', 'REQ', 'OPT'],
        example=['Contracts', 'CTR-001', 'Contains', 'Hardware', 'HW-001', 'sample row'],
        tab_color=TAB_CUSTOM,
        dropdowns={1: 'Module', 3: 'RelationshipType', 4: 'Module'},
    )

    add_catalogs_sheet(wb)

    out_path = os.path.join(OUT_DIR, 'OPRIVA_IMPORT_TEMPLATE_CANONICAL.xlsx')
    wb.save(out_path)
    return out_path


def main():
    if not os.path.isdir(OUT_DIR):
        os.makedirs(OUT_DIR)
    paths = [
        build_msp_template(),
        build_internal_it_template(),
        build_canonical_template(),
    ]
    for p in paths:
        size = os.path.getsize(p)
        print(f'Wrote {p} ({size} bytes)')


if __name__ == '__main__':
    main()
